# -*- coding: utf-8 -*-
import pymc3 as pm
import numpy as np
import pandas as pd
from theano import shared
import scipy.stats as stats
import matplotlib.pyplot as plt
import arviz as az
import re
from datetime import datetime
import openpyxl


az.style.use('arviz-darkgrid')

np.random.seed(1)

# In[] Загрузка данных
file_name = 'Ошибки моделей\\Ретроспектива\\Несколько моделей (внутри train).xlsx'
sw_hist_excel = pd.ExcelFile(file_name)

file_name = 'Ошибки моделей\\Ретроспектива\\Несколько моделей для ансамбля.xlsx'
sw_excel = pd.ExcelFile(file_name)

file_name = 'Ошибки моделей\\Ретроспектива\\xgb_hist.xlsx'
xgb_hist_excel = pd.ExcelFile(file_name)

file_name = 'Ошибки моделей\\Ретроспектива\\xgb.xlsx'
xgb_excel = pd.ExcelFile(file_name)

file_name = 'Ошибки моделей\\Ретроспектива\\eln_hist.xlsx'
eln_hist_excel = pd.ExcelFile(file_name)

file_name = 'Ошибки моделей\\Ретроспектива\\eln.xlsx'
eln_excel = pd.ExcelFile(file_name)

file_name = 'Ошибки моделей\\Ретроспектива\\CRM.xlsx'
crm_excel = pd.ExcelFile(file_name)

file_name = 'Ошибки моделей\\Ретроспектива\\CRM_hist.xlsx'
crm_hist_excel = pd.ExcelFile(file_name)

wb = openpyxl.Workbook()    # создаем новый excel-файл

for index_, well_num in enumerate(sw_excel.sheet_names[:-1]):
# for index_, well_num in enumerate(sw_excel.sheet_names[-7:-1]):
    try:
        if well_num in ['32', '22', '7', '55']:
            continue
        sw_hist_sheet = pd.read_excel(sw_hist_excel, sheet_name=well_num)
        sw_sheet = pd.read_excel(sw_excel, sheet_name=well_num)
        xgb_sheet = pd.read_excel(xgb_excel, sheet_name=well_num)
        xgb_hist_sheet = pd.read_excel(xgb_hist_excel, sheet_name=well_num)
        eln_sheet = pd.read_excel(eln_excel, sheet_name=well_num)
        eln_hist_sheet = pd.read_excel(eln_hist_excel, sheet_name=well_num)
        crm_sheet = pd.read_excel(crm_excel, 0)
        crm_hist_sheet = pd.read_excel(crm_hist_excel, 0)
        
        # In[] Формирование обучающей выборки
        time = pd.to_datetime(sw_hist_sheet.iloc[[index for index,row in sw_hist_sheet.iterrows() if row.iat[0] >= datetime(2019, 11, 1)], 0])
        k = len(time)
        n = len(sw_hist_sheet.iloc[:, 0])
        k = n - k
        time_zero = time.values[-1]
        a = list(reversed([abs((elem - time_zero).days) for elem in time]))
        time_deltas = shared(np.array(list(reversed([abs((elem - time_zero).days) for elem in time]))))

        Q_real=shared(sw_hist_sheet.iloc[k:n,1].values)
        Q_model1=shared(sw_hist_sheet.iloc[k:n,2].values)
        Q_model2=shared(sw_hist_sheet.iloc[k:n,3].values)
        Q_model3=shared(sw_hist_sheet.iloc[k:n,4].values)
        Q_model4=shared(sw_hist_sheet.iloc[k:n,5].values)
        crm_model = shared(crm_hist_sheet.loc[:, well_num].values)
        xgb_model = shared(xgb_hist_sheet.iloc[[index for index,row in xgb_hist_sheet.iterrows() if row.iat[7] in list(time)], 8].values)
        eln_model = shared(eln_hist_sheet.iloc[[index for index,row in eln_hist_sheet.iterrows() if row.iat[7] in list(time)], 8].values)

        # In[] Настройка модели
        with pm.Model() as model_g:
            w0 = pm.Normal('w0', mu=0, sd=1)
            w1 = pm.Normal('w1', mu=0, sd=1)
            w2 = pm.Normal('w2', mu=0, sd=1)
            w3 = pm.Normal('w3', mu=0, sd=1)
            w4 = pm.Normal('w4', mu=0, sd=1)
            w_crm = pm.Normal('w_crm', mu=0, sd=1)
            w_xgb = pm.Normal('w_xgb', mu=0, sd=1)
            w_eln = pm.Normal('w_eln', mu=0, sd=1)
            # w6 = pm.HalfNormal('w6', sd=80)
            # w7 = pm.HalfNormal('w7', sd=30)

            mu = pm.Deterministic('mu',
                                  w0 + w1 * Q_model1 + w2 * Q_model2 + w3 * Q_model3 + w4 * Q_model4
                                  + w_crm * crm_model
                                  + w_xgb * xgb_model
                                  + w_eln * eln_model
                                  )
            sd = pm.Deterministic('sd', w6 + w7 * time_deltas)
            y_pred = pm.Normal('y_pred', mu=mu, sd=sd, observed=Q_real)

            trace_g = pm.sample(1500, tune=1000, target_accept=.95, cores=-1)

        # Отображение статистики по параметрами. Хорошо, если распределения гладкие, унимодальные
        # az.plot_trace(trace_g, var_names=['w0', 'w1', 'w2', 'w3', 'w4', 'w_crm', 'w6', 'sd'])

        # In[] Проверка модели на периоде обучения
        # Генерация сэмплов прогнозов из построенной модели. Нужно для того, чтобы было много сэмплов для построения доверительного интервала
        ppc = pm.sample_posterior_predictive(trace_g, samples=700, model=model_g)
        # In[]
        # Отрисовка прогноза наиболее вероятного(среднего) прогноза
        w0_m = trace_g['w0'].mean()
        w1_m = trace_g['w1'].mean()
        w2_m = trace_g['w2'].mean()
        w3_m = trace_g['w3'].mean()
        w4_m = trace_g['w4'].mean()
        w_crm_m = trace_g['w_crm'].mean()
        w_xgb_m = trace_g['w_xgb'].mean()
        w_eln_m = trace_g['w_eln'].mean()
        w6_m = trace_g['w6'].mean()
        w7_m = trace_g['w7'].mean()


        plt.plot(time, w0_m + w1_m * Q_model1.get_value() + w2_m * Q_model2.get_value() + w3_m * Q_model3.get_value()
                 + w4_m * Q_model4.get_value()
                 + w_crm_m * crm_model.get_value()
                 + w_xgb_m * xgb_model.get_value()
                 + w_eln_m * eln_model.get_value()
                 ,c='k', label='Q-наиболее вероятное')

        plt.plot(time, Q_real.get_value(), 'C0.')
        plt.plot(time, Q_model1.get_value(), 'g--', label='Q_модель1')
        plt.plot(time, Q_model2.get_value(), 'b--', label='Q_модель2')
        plt.plot(time, Q_model3.get_value(), 'y--', label='Q_модель3')
        plt.plot(time, Q_model4.get_value(), 'm--', label='Q_модель4')
        plt.plot(time, crm_model.get_value(), 'c--', label='CRM')
        plt.plot(time, xgb_model.get_value(), 'r--', label='XGBoost')
        plt.plot(time, eln_model.get_value(), color='navy', linestyle='--', label='ElasticNet')

        # Отрисовка доверительных интервалов уровнем доверия 50% и 95%
        az.plot_hpd(time, ppc['y_pred'], credible_interval=0.5, color='tan', smooth=False)
        az.plot_hpd(time, ppc['y_pred'], credible_interval=0.95, color='gray', smooth=False)
        plt.rc('legend', fontsize=8)
        plt.rc('axes', labelsize=8)  # fontsize of the tick labels
        plt.title('Адаптация модели')
        plt.legend()
        plt.xticks(rotation=25)
        plt.savefig(str(well_num) + ' - обучение - ретроспектива.png', dpi=200, bbox_inches='tight')
        plt.tight_layout()
        plt.show()

        # In[] Построение модели на периоде прогноза
        time = pd.to_datetime(sw_sheet.iloc[[index for index,row in sw_sheet.iterrows() if row.iat[0] >= datetime(2020, 2, 1)], 0])
        k = len(time)
        n = len(sw_sheet.iloc[:, 0])
        k = n - k
        # Переопределение входных параметров модели
        Q_model1.set_value(sw_sheet.iloc[k:, 2])
        Q_model2.set_value(sw_sheet.iloc[k:, 3])
        Q_model3.set_value(sw_sheet.iloc[k:, 4])
        Q_model4.set_value(sw_sheet.iloc[k:, 5])
        crm_model.set_value(crm_sheet.loc[:, well_num].values)
        xgb_model.set_value(
            xgb_sheet.iloc[[index for index, row in xgb_sheet.iterrows() if row.iat[7] in list(time)], 8].values)
        eln_model.set_value(
            eln_sheet.iloc[[index for index, row in eln_sheet.iterrows() if row.iat[7] in list(time)], 8].values)
        time_deltas.set_value(np.array([abs((elem - time_zero).days) for elem in time]))

        # In[]
        # Оценка сдвига последней точки факта от прогноза. Это сдвиг будет использован для сдвига будущего прогноза
        delta = sw_sheet.iloc[k - 1, 1] - (
                w0_m + w1_m * sw_sheet.iloc[k - 1, 2] + w2_m * sw_sheet.iloc[k - 1, 3] + w3_m *
                sw_sheet.iloc[k - 1, 4]
                + w4_m * sw_sheet.iloc[k - 1, 5]
                + w_crm_m * crm_sheet.loc[0, well_num]
                + w_xgb_m * xgb_sheet.iloc[
                    [index for index, row in xgb_sheet.iterrows() if row.iat[2] == sw_sheet.iloc[k - 1, 0]], 3].values[0]
                + w_eln_m * eln_sheet.iloc[
                    [index for index, row in eln_sheet.iterrows() if row.iat[2] == sw_sheet.iloc[k - 1, 0]], 3].values[0])

        # In[]
        # Генерация сэмплов прогнозов из построенной модели. Нужно для того, чтобы было много сэмплов для построения доверительного интервала
        ppc2 = pm.sample_posterior_predictive(trace_g, samples=1000, model=model_g)


        plt.plot(time, delta + w0_m + w1_m * Q_model1.get_value() + w2_m * Q_model2.get_value() +
                 w3_m * Q_model3.get_value() + w4_m * Q_model4.get_value()
                 + w_crm_m * crm_model.get_value()
                 + w_xgb_m * xgb_model.get_value()
                 + w_eln_m * eln_model.get_value()
                 ,c='k', label='Q-наиболее вероятное')

        plt.plot(time, sw_sheet.iloc[k:, 1].values, 'C0.')
        plt.plot(time, Q_model1.get_value(), 'g--', label='Q_модель1')
        plt.plot(time, Q_model2.get_value(), 'b--', label='Q_модель2')
        plt.plot(time, Q_model3.get_value(), 'y--', label='Q_модель3')
        plt.plot(time, Q_model4.get_value(), 'm--', label='Q_модель4')
        plt.plot(time, crm_model.get_value(), 'c--', label='CRM')
        plt.plot(time, xgb_model.get_value(), 'r--', label='XGBoost')
        plt.plot(time, eln_model.get_value(), color='navy', linestyle='--', label='ElasticNet')

        # Отрисовка доверительных интервалов уровнем доверия 50% и 95%
        az.plot_hpd(time, ppc2['y_pred'] + delta, credible_interval=0.5, color='tan', smooth=False)
        az.plot_hpd(time, ppc2['y_pred'] + delta, credible_interval=0.95, color='gray', smooth=False)
        plt.rc('legend', fontsize=8)
        plt.rc('axes', labelsize=8)  # fontsize of the tick labels

        plt.title('Прогноза модели')
        plt.legend()
        plt.tight_layout()
        plt.xticks(rotation=30)
        plt.savefig(str(well_num) + ' - прогноз - ретроспектива.png', dpi=200, bbox_inches='tight')
        plt.show()

        Q_mean = delta + w0_m + w1_m * Q_model1.get_value() + w2_m * Q_model2.get_value()\
                 + w3_m * Q_model3.get_value() + w4_m * Q_model4.get_value() \
                 + w_crm_m * crm_model.get_value() \
                 + w_xgb_m * xgb_model.get_value() \
                 + w_eln_m * eln_model.get_value()
        # добавляем новый лист
        wb.create_sheet(title=well_num, index=index_)

        # получаем лист, с которым будем работать
        sheet_abs = wb[well_num]

        cell_abs = sheet_abs.cell(row=1, column=1)
        cell_abs.value = 'Q_liq ансамбль'

        for row in range(2, len(Q_mean) + 2):
            cell_abs = sheet_abs.cell(row=row, column=1)
            cell_abs.value = Q_mean[row - 2]
    except:
        # raise TypeError
        continue

wb.save('Ошибки моделей\\Ансамбль прогноз ретроспективный + CRM.xlsx')


